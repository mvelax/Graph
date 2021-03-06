import xlsxwriter
import csv
import glob
from xlsxwriter.utility import xl_col_to_name
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import load_workbook
import sys
import time
from time import mktime
import datetime
from datetime import date
import re
#Read a csv file and create graphs
"""
To add a new graph:
1) Add a new entry with a new number in the COMMON_GRAPHS 
	or NOKIA_GRAPHS list.
2) Add a new chart_column_pair entry in the addANRExecChart
	method. Here you determine which column will be used
	to set the height of the execution chart bars.
	Ideally it will be one of the numbers you specified
	in step one.
"""


def addDefaultColumnSeries(chart, sheet_name, y2_axis, column, last_row):
	STARTING_ROW = 1
	HEADER = 0
	"""
	Add a series using the header of that column,
	values will use row 2 to the one determined in rows.
	leftmost column will determine the category.
	chart: chart object where series will be added.
	sheet_name: string containing the sheet name.
	y2_axis: boolean, states if series should be part of y_axis 1(False) or 2(True).
	column: 0 index numer indicating the columnd witht the desired data.
	last_row: number indicating how many 
	"""
	if not y2_axis:
		chart.add_series({
			"name":			[sheet_name, HEADER,column],
			"categories":	[sheet_name, STARTING_ROW, HEADER, last_row, HEADER],
			"values":		[sheet_name, STARTING_ROW, column, last_row,column]
		})
	else:
		chart.add_series({
			"name":			[sheet_name, HEADER,column],
			"categories":	[sheet_name, STARTING_ROW, HEADER, last_row, HEADER],
			"values":		[sheet_name, STARTING_ROW, column, last_row,column],
			"y2_axis": 1
		})
		
	
	
def makeChart(workbook, data_worksheet, info, last_row, anr_sheet):
	"""
	Creates a chart in the worksheet, inspired in grafico1.
	info is a tuple like:
	([y1_axis_cols],[y2_axis_cols],number,name)
	"""
	chart_sheet = workbook.add_chartsheet("Grafico{}".format(info[2]))
	chart = workbook.add_chart({"type":"line"})
	addANRExecChart(workbook, chart, info[2], anr_sheet.get_name(), data_worksheet.get_name(), last_row)
	for column in info[0]:
		addDefaultColumnSeries(chart, data_worksheet.get_name(), False, column, last_row)
	for column in info[1]:
		addDefaultColumnSeries(chart, data_worksheet.get_name(), True, column, last_row)
	chart.set_legend({'position': 'bottom'})
	chart.set_title({"name":info[3]})
	chart_sheet.set_chart(chart)
	
	
def addANRExecChart(workbook, chart, chart_number, anr_sheet_name, data_worksheet_name, last_row):
	"""
	Create the ANR execution chart and combine the main chart with it.
	It will read the correct column in the ANR helper table,
	to determine the correct height.
	The column to be read is determined by the chart_number and must be hard coded.
	"""
	ANR_EXECUTION_SERIES_NAME = "ANR Execution"
	STARTING_ROW = 1 
	HEADER = 0
	#Using the chart number, we can determine which row from ANR helper to read.
	chart_column_pair = {1:2, 2:5, 3:10, 4:8, 5:15, 6:17, 8:2, 9:2}
	anr_chart = workbook.add_chart({"type":"column"})
	anr_chart.add_series({
			"name":			ANR_EXECUTION_SERIES_NAME,
			"categories":	[data_worksheet_name, STARTING_ROW, HEADER, last_row, HEADER],
			"values":		[anr_sheet_name, STARTING_ROW, chart_column_pair[chart_number], last_row, chart_column_pair[chart_number]],
			"fill":			{"color": "#fbbc05"},
			"gap":			0
		})
	chart.combine(anr_chart)
	
def makeANRHelperTable(anr_sheet, header, total_rows, worksheet_name):
	#Note that the number of rows passed in is 0 indexed.
	#excel_row is 1 indexed.
	columns = len(header)
	ANR_bar_value = "=('Helper Table'!$C{excel_row})*MAX('{worksheet_name}'!{excel_col}$2:{excel_col}${total_rows})"
	anr_sheet.write_row(0,0,header)
	for row in range(total_rows):
		for col in range(columns):
			anr_sheet.write(row+1, 
							col, 
							ANR_bar_value.format(
												excel_row = str(row+1+1),
												excel_col = xl_col_to_name(col),
												total_rows = str(total_rows),
												worksheet_name=worksheet_name
							)
			)
	
def parseExecutionCalendar():
	"""
	Returns a dictionary based on the execution calendar.
	The execution calendar must follow convention.
	Dates and headers in third row.
	"ANR", "LMS" and "LMS/ANR" are the only accepted values.
	
	The dictionary looks like:
	{
		RNC_NAME:{
					ANR:set([date objects])
					LMS:set([date objects])
				}
		...
	
	}
	The sets include all dates when ANR or LMS was executed 
	for a given RNC.
	"""
	wb = load_workbook("Calendario edenrock Historico Closed Loop.xlsx")
	ws = wb.active

	exec_db = {}
	rnc_col = 0
	date_col = -1

	for i,header_cell in enumerate(ws.rows[2]):
		#print type(header_cell.value)
		if header_cell.value == 'RNC':
			rnc_col = i
		elif type(header_cell.value) is datetime.datetime and date_col == -1:
			date_col = i

	dates = ws.rows[2][date_col:]
		

	for row in ws.rows[3:]:
		rnc_name = row[rnc_col].value
		exec_db[rnc_name] = {"ANR":set(),"LMS":set()}
		for date_var,col in zip(dates,row[date_col:]):
			#print col.value
			if col.value != None:
				upper_val = col.value.upper()
				date_val = date_var.value.date()
				if upper_val == "ANR":
					exec_db[rnc_name]["ANR"].add(date_val)
				elif upper_val == "LMS":
					exec_db[rnc_name]["LMS"].add(date_val)
				elif upper_val == "LMS/ANR":
					exec_db[rnc_name]["ANR"].add(date_val)
					exec_db[rnc_name]["LMS"].add(date_val)
					
	return exec_db
	
	
def main():
	if len(sys.argv) != 2 or sys.argv[1] == "-h":
		print "Make sure that file 'Calendario edenrock Historico Closed Loop.xlsx' is in the same directory."
		print "Correct usage: python createGraph.py  <vendor(nokia or ericsson)>"
		quit()
	#Use user input
	#input_file = sys.argv[1]
	vendor = sys.argv[1]
	INPUT = "kpi_checker*.csv"

	if vendor != "nokia" and vendor != "ericsson":
		print "Not a valid vendor: nokia or ericsson"
		quit()


	csv_file_list = glob.glob(INPUT)
	COMMON_GRAPHS = [
				#([y1_axis_cols],[y2_axis_cols],number,name)
				([2],[4,14],1,"Volumen de trafico de voz cursado & Tasa de caidas de voz & Tasa de fallos de accesibilidad de voz"),
				([5],[6,12],2,"Volumen de trafico de datos & Tasa de fallos de accesibilidad & Tasa de caidas de datos"),
				([12,10],[],3,"Tasa de Accesibilidad HSDPA & Tasa Accesibilidad HSUPA"),
				([8],[],4,"Tasa de llamadas de voz originadas en 3G y que terminan en 2G"),
				([15],[16],5,"Volumen de SHO y Tasa de exito de SHO"),
				([17],[18],6,"Volumen de IFHO y Tasa de Fallos de IFHO")
			]
			
	NOKIA_GRAPHS = [
				([2],[21],8,"Volumen de Voz & Total de caidas por detected"),
				([2],[22],9,"Volumen de Voz & Tasa de Caidas por Detected")
			]
	OFFSET = 2 #Difference between columns in Metricas_Datos and Helper Table.
	SUMMARY_HEADER = ["", "Antes", "Despues", "% Diferencia"]	
	SUMMARY_TABLE_ORIGIN = (12,1)
	SUMMARY_SHEET_NAME = "Resumen"
	ANR_SHEET_NAME = "ANR Helper"


	for file_name in csv_file_list:
		CSV_TIME_FORMAT = "%Y-%m-%d %H:%M:%S %Z"
		RNC_NAME_RE_STRING = r"^.*[0-9]_(.*)\.csv"
		rnc_name_re = re.compile(RNC_NAME_RE_STRING)
		
		file_name_split = file_name.split(".")
		assert(len(file_name_split) == 2)
		file_name_no_ext = file_name_split[0]

		wb = xlsxwriter.Workbook(file_name_no_ext+".xlsx", {'strings_to_numbers':  True})
		ws = wb.add_worksheet("Metricas_Datos") #Holds csv data
		helper_sheet = wb.add_worksheet("Helper Table") #Calculates pre and post ANR averages.
		anr_sheet = wb.add_worksheet(ANR_SHEET_NAME) #Holds values for ANR exec graphs
		
		#Formats
		good = wb.add_format()
		good.set_bg_color("#afceb8")
		good.set_font_color("#27a700")
		
		bad = wb.add_format()
		bad.set_bg_color("#f6918c")
		bad.set_font_color("#c50900")
		
		summary_kpi_format = wb.add_format()
		summary_kpi_format.set_bg_color('#759FCD')
		
		green_bg = wb.add_format()
		green_bg.set_bg_color('#1e6810')
		
		percent = wb.add_format()
		percent.set_num_format(10)   
		
		#Conditional formats
		greater_good = {'type':     'cell',
						'criteria': '>=',
						'value':    0,
						'format':   good}
		greater_bad = {'type':     'cell',
						'criteria': '>=',
						'value':    0,
						'format':   bad}
		lower_good = {'type':     'cell',
						'criteria': '<',
						'value':    0,
						'format':   good}
		lower_bad = {'type':     'cell',
						'criteria': '<',
						'value':    0,
						'format':   bad}
		

		#Need to read the csv dates to look them up in the execution matrix.
		time_col = []
		
		#Open the csv, read it and write it in data_worksheet
		#Also, while reading memorize the dates.
		with open(file_name, "r") as csv_file:
			csv_reader = csv.reader(csv_file)
			number_of_rows = 0
			header = []

			for i,row in enumerate(csv_reader):
				try:
					time_col.append(datetime.datetime.strptime(row[0],CSV_TIME_FORMAT))
				except ValueError:
					time_col.append(row[0])
				if i==0:
					header = row
				clean_row = []
				for element in row:
					if element=="None":
						clean_row.append("")
					else:
						clean_row.append(element)
				ws.write_row(i,0,clean_row)
			else:
				number_of_rows = i+1
				number_of_columns = len(row)
				helper_sheet.write_column(0,0,time_col)
			
			#Copy from input file into helper sheet.
			#Make them date objects so they are compatible with dict.
			date_col = [x.date() for x in time_col[1:]]
			#Get the RNC name.
			re_match_filename = rnc_name_re.match(file_name)
			rnc_name = re_match_filename.groups()[0]
			#Get the execution dictionary
			exec_dict = parseExecutionCalendar()
			average_flag = 0 #Turns into 1 after first ANR
			for i,date_var in enumerate(date_col):
				anr_exec = ""
				lms_exec = ""
				if date_var in exec_dict[rnc_name]["ANR"]:
					anr_exec = 1
					average_flag = 1
				if date_var in exec_dict[rnc_name]["LMS"]:
					lms_exec = 1
					
				row = [lms_exec, anr_exec, average_flag]
				helper_sheet.write_row(i+1,1,row)
				
			
			"""
			with open(input_file, "r") as anr_file:
				anr_reader = csv.reader(anr_file)
				for i,row in enumerate(anr_reader):
					if i==0:
						continue
					helper_sheet.write_row(i,1,row)
			"""
			#Make the Helper table
			helper_header = [header[0]] + ["ANR Execution", "ANR Execution", "ANR avg"] + header[2:]
			helper_sheet.write_row(0,0,helper_header)
			avg_if_formula = "=AVERAGEIF($D:$D,$D{0},Metricas_Datos!{1}:{1})"
			for row in range(1,number_of_rows):
				for col in range(2,number_of_columns):
					col_letter = xl_col_to_name(col)
					helper_sheet.write(row,col+OFFSET,avg_if_formula.format(row+1,col_letter))
			
			#Make ANR helper table
			makeANRHelperTable(anr_sheet, header, number_of_rows, ws.get_name())
			
			#Create the summary table.
			summary_sheet = wb.add_worksheet(SUMMARY_SHEET_NAME)
			summary_sheet.write_row(SUMMARY_TABLE_ORIGIN[0], SUMMARY_TABLE_ORIGIN[1], SUMMARY_HEADER, green_bg)
			difference_cell = "=('{0}'!{1}-'{0}'!{2})/'{0}'!{2}"
			copy_cell = "='{}'!{}"
			for x in range(number_of_columns-2):
				#Using magic numbers to copy the summary table!
				summary_sheet.write(SUMMARY_TABLE_ORIGIN[0]+1+x,
									SUMMARY_TABLE_ORIGIN[1],
									copy_cell.format(helper_sheet.get_name(),xl_rowcol_to_cell(0,4+x)),
									summary_kpi_format)
				summary_sheet.write(SUMMARY_TABLE_ORIGIN[0]+1+x,
									SUMMARY_TABLE_ORIGIN[1]+1,
									copy_cell.format(helper_sheet.get_name(),xl_rowcol_to_cell(1,4+x)))
				summary_sheet.write(SUMMARY_TABLE_ORIGIN[0]+1+x,
									SUMMARY_TABLE_ORIGIN[1]+2,
									copy_cell.format(helper_sheet.get_name(),xl_rowcol_to_cell(number_of_rows-1,4+x)))
				before_cell = xl_rowcol_to_cell(SUMMARY_TABLE_ORIGIN[0]+x+1,SUMMARY_TABLE_ORIGIN[1]+1)
				after_cell = xl_rowcol_to_cell(SUMMARY_TABLE_ORIGIN[0]+x+1,SUMMARY_TABLE_ORIGIN[1]+2)
				summary_sheet.write(SUMMARY_TABLE_ORIGIN[0]+1+x,
									SUMMARY_TABLE_ORIGIN[1]+3,
									difference_cell.format(summary_sheet.get_name(),after_cell,before_cell),
									percent)
			
			#Conditional formats
			good_metrics = []
			bad_metrics = []
			if vendor == "nokia":
				good_metrics = [12,14,16,18]
				bad_metrics = [1,2,4,5,6,7,8,9,10,11,19]
			else:
				bad_metrics = [1,2,4,5,6,7,8,9,10,11,12,16]
				good_metrics = [14]
				
			for row in good_metrics:
				summary_sheet.conditional_format(SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3,SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3, greater_good)
				summary_sheet.conditional_format(SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3,SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3, lower_bad)

			for row in bad_metrics:
				summary_sheet.conditional_format(SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3,SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3, greater_bad)
				summary_sheet.conditional_format(SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3,SUMMARY_TABLE_ORIGIN[0]+row+1, SUMMARY_TABLE_ORIGIN[1]+3, lower_good)
					
			#Magic numbers set column width. (Empirical)
			summary_sheet.set_column(1,1,54.57)
			summary_sheet.set_column(4,4,11.43)
			summary_sheet.activate()	

			#Make all charts
			for info in COMMON_GRAPHS:
				makeChart(wb,ws,info,number_of_rows, anr_sheet)
			if vendor == "nokia":
				for info in NOKIA_GRAPHS:
					makeChart(wb,ws,info,number_of_rows, anr_sheet)
			
		wb.close()
	
if __name__ == "__main__":
	main()